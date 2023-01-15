#!/usr/bin/env python3.10
import random
import matplotlib.pyplot as grafico
from getpass_asterisk.getpass_asterisk import getpass_asterisk
from openpyxl import Workbook, load_workbook
wb = load_workbook('/Users/sofi/Desktop/REBECA-MASTER/PYTHON/PYTHON BÁSICO/PRÁCTICA/Excel_respuestas.xlsx')
ws = wb.active


def elegir_modo():
    respuesta = 0

    modos = ['Solitario', '2 jugadores', 'Estadística', 'Salir']

    while(respuesta < 1 or respuesta > (len(modos) + 1)):
        for index, mode in enumerate(modos):
            print(f'{index + 1} . {mode}')

        respuesta = int(input())

        if respuesta > (len(modos) + 1):
            print('Por favor, escriba una respuesta valida')
    return respuesta
def solitario():
    respuesta_dificultad = 0
    dificultades =['Facil','Medio', 'Difícil']
    intentos_dificultad = [20, 12, 5]

    while (respuesta_dificultad < 1 or respuesta_dificultad> (len(dificultades)+1)):
        for index, dificultad in enumerate(dificultades):
            print(f'{index + 1} . {dificultad}')

        respuesta_dificultad = int(input())
        if (respuesta_dificultad < 1 or respuesta_dificultad > (len(dificultades)+1)):
            print('Por favor, escriba una respuesta valida')

    print(f'Has eligido el juego {dificultades[respuesta_dificultad -1]}, tienes {intentos_dificultad[respuesta_dificultad -1]} intentos para adivinar el número comprendido entre 1 y 15.')

    aleatorio = random.randint(1, 15)
    intentos = 0
    max_intentos = intentos_dificultad[respuesta_dificultad -1]
    print('Es la hora, intenta adivinar el número!')
    while intentos <= max_intentos:
        numero = int(input())
        if numero == aleatorio:
            print('ENHORABUENA, HAS ACERTADO!')
            print('Vuelve a introducir tu nombre por favor')
            ws.append([input(),'GANADO'])
            wb.save('/Users/sofi/Desktop/REBECA-MASTER/PYTHON/PYTHON BÁSICO/PRÁCTICA/Excel_respuestas.xlsx')
            break

        if numero < aleatorio:
            print('El número que buscas es mayor')
            intentos += 1
            print(f'Llevas {intentos} intentos')
        if numero > aleatorio :
            print('El número que buscas es menor')
            intentos += 1
            print(f'Llevas {intentos} intentos')

        if intentos > max_intentos:
            print('Ohh no has tenido suerte esta vez!, vuelve a intentarlo.')
            print('Vuelve a introducir tu nombre por favor.')
            ws.append([input(), 'PERDIDO'])
            wb.save('/Users/sofi/Desktop/REBECA-MASTER/PYTHON/PYTHON BÁSICO/PRÁCTICA/Excel_respuestas.xlsx')

    return respuesta_dificultad


def dos_jugadores():
    respuesta_dificultad = 0
    dificultades = ['Facil', 'Medio', 'Difícil']
    intentos_dificultad=[20,12,5]


    while (respuesta_dificultad < 1 or respuesta_dificultad > (len(dificultades) + 1)):
        for index, dificultad in enumerate(dificultades):
            print(f'{index + 1} . {dificultad}')

        respuesta_dificultad = int(input())
        if (respuesta_dificultad < 1 or respuesta_dificultad > (len(dificultades) + 1)):
            print('Por favor, escriba una respuesta valida')

    print(f'Has eligido el juego {dificultades[respuesta_dificultad - 1]} introduce un número entre 1 y 1000 para que el jugador 2 lo adivine.')
    intentos = 0
    max_intentos = intentos_dificultad[respuesta_dificultad -1]

    adivinar = int(getpass_asterisk("Introduce tu numero: "))
    print(f'Es tu momento jugador 2, tienes {intentos_dificultad[respuesta_dificultad-1]} intentos para adivinar el numero elegido por el jugador 1.El número se encuentra entre 1 y 1000. Iremos dándote pistas de si el número buscado es mayor o menor. VAMOS A POR ELLO!')

    while intentos <= max_intentos:
        numero = int(input())
        if numero == adivinar:
            print('ENHORABUENA, HAS ACERTADO!')
            print('Vuelve a introducir tu nombre por favor')
            ws.append([input(), 'GANADO'])
            wb.save('/Users/sofi/Desktop/REBECA-MASTER/PYTHON/PYTHON BÁSICO/PRÁCTICA/Excel_respuestas.xlsx')
            break
        if numero < adivinar:
            print('El número que buscas es mayor')
            intentos += 1
            print(f'Llevas {intentos} intentos')

        if numero > adivinar:
            print('El número que buscas es menor')
            intentos += 1
            print(f'Llevas {intentos} intentos')

        if intentos > max_intentos :
            print('Ohh no has acertado el número, vuelve a intentarlo!')
            print('Escribe tu nombre de nuevo por favor')
            ws.append([input(), 'PERDIDO'])
            wb.save('/Users/sofi/Desktop/REBECA-MASTER/PYTHON/PYTHON BÁSICO/PRÁCTICA/Excel_respuestas.xlsx')

    return respuesta_dificultad

def estadistica():
    max_row = ws.max_row
    ganado = 0
    perdido = 0
    for name in range(1, max_row + 1):
        print(ws[name][0].value)
        print(ws[name][1].value)
        print('-----')
        if ws[name][1].value == 'GANADO':
            ganado += 1
        if ws[name][1].value == 'PERDIDO':
            perdido += 1

    print(f'Han ganado {ganado} y perdido {perdido} ')

    grafico.title('Resumen respuestas')
    grafico.xlabel('Total ganado o perdido')
    grafico.ylabel('Nº GANADO O PERDIDO')

    x = ['Ganado', 'Perdido']
    y = [ganado, perdido]
    grafico.bar(x, y)
    grafico.show()
    return





def salir():
    print('Hasta pronto!')



def funcion_principal():
    print('Bienvenid@ a este maravilloso juego de adivinanzas. Aquí podrás demostrar tus habilidades. ')
    print('En primer lugar, necesito que escribas tu nombre para empezar a jugar.')
    nombre = input()
    print(f'Hola! {nombre} te vas introducir en un juego donde todas las respuestas son validas pero no correctas. Para seguir avanzando en el juego, es importante que te concentres y des con el número correcto!')

    print('Elige el modo del juego:')
    modo_elegido = elegir_modo()


    match modo_elegido:
        case 1:
            print('Has eligido el modo SOLITARIO, a continuación se generará un número aleatorio entre el 1 y 15, tienes que adivinarlo. Pero primero eligue el tipo de dificultad.')
            solitario()
        case 2:
            print('Has elegido el modo ¡2 JUGADORES!. Este juego consiste en que el jugador 2 acierte el número introducido por el jugador 1.')
            dos_jugadores()
        case 3:
            print('Has elegido el modo Estadística. A continuación se mostrará el resumen total de jugadores que han jugado hasta el momento.')
            estadistica()
        case 4:
            print('Has acabo el juego.')
            salir()

    while True:
        continuar = input('Quieres volver a jugar? S / N :')

        if continuar.lower() in "s si y yes":
            return funcion_principal()

        else:
            print('El juego ha terminado, ¡hasta pronto!')
            break


funcion_principal()




